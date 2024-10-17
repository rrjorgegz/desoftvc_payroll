from odoo import fields, models, api, _
from odoo.exceptions import UserError
import xlrd
import tempfile
import binascii
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from dateutil.relativedelta import relativedelta

class ImportarVentasWiz(models.TransientModel):
    _name = "comercializador.importar.ventas.wiz"

    import_file = fields.Binary('Import File', required = True)
    filename = fields.Char("File Name")
    division_id = fields.Many2one(comodel_name='comercializador.division', required=True, string="Division vendedora", tracking=True, ondelete='cascade')
    producto_id = fields.Many2one(comodel_name='comercializador.producto', required=True, string="Producto", tracking=True, ondelete='cascade')
    

    def button_import_xlsx_sales(self):
        try:
            fp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            fp.write(binascii.a2b_base64(self.import_file))
            fp.seek(0)
            fp.close
        except:
            raise UserError('El archivo no es válido.')

        try:
            workbook = xlrd.open_workbook(fp.name, on_demand=True)
            sheet = workbook.sheet_by_index(0)
            
            wb = load_workbook(fp.name)
            
        except Exception as e:
            raise UserError('El archivo no es válido.')
        
        if sheet.ncols == 0:
            return
        
        clientes = self.env['comercializador.cliente'].sudo().search([])
        
        import_lines = []
        new_clients_ids = []
        
        for row in range(2, sheet.nrows):
            line = {}
            for col in range(5):
                if col == 0:
                    line['fecha_venta'] = sheet.cell_value(row, col)
                elif col == 1:
                    line['factura'] = sheet.cell_value(row, col)
                elif col == 2:
                    line['importe_venta'] = sheet.cell_value(row, col)
                elif col == 3:
                    line['nombre_cliente'] = sheet.cell_value(row, col)
                elif col == 4:
                    line['codigo_concepto_facturacion'] = sheet.cell_value(row, col)
            
            cliente_id = 0      
            for cliente in clientes:
                if cliente.nombre == line['nombre_cliente']:
                    cliente_id = cliente.id
                    
            
            if cliente_id == 0:
                importe = float(line['importe_venta'])
                if importe >= 370 and importe <= 371:
                    nuevo_cliente_id = self.env['comercializador.cliente'].create({
                        'nombre': line['nombre_cliente'],
                        'provincia': 'villa_clara',
                    }).id
                    new_clients_ids.append(nuevo_cliente_id)
                    line['cliente_id'] = nuevo_cliente_id
                else: # dejo en 0 el id del cliente para crearlo en el wizard de procesar ventas
                    line['cliente_id'] = cliente_id
            else:
                line['cliente_id'] = cliente_id
                
            import_lines.append(line)
            
        ventas_ids = []
        ventas_mayores = []

        for import_line in import_lines:
            venta = self.env['comercializador.venta'].sudo().search([('factura','=',import_line.get('factura'))])
            if not venta: # si no se ha registrado esta venta todavia , asumiento que la factura es unica en el versat
                importe = float(import_line.get('importe_venta'))
                if importe >= 370 and importe <= 371: # si el importe de la venta es 370$ el periodo de subbscripcion es de 1 mes
                    venta_id = self.env['comercializador.venta'].create({
                            'fecha_venta': datetime.strptime(import_line.get('fecha_venta').strip(), '%d/%m/%Y').date(),
                            'importe_venta': importe,
                            'tipo_venta': 'subscripcion',
                            'cliente_id': import_line.get('cliente_id'),
                            'factura': import_line.get('factura'),
                            'codigo_concepto_facturacion': import_line.get('codigo_concepto_facturacion'),
                            'division_id': self.division_id.id,
                            'producto_id': self.producto_id.id,
                        }).id # la venta la creo aqui solo si el importe de suscripcion equivale a 1 mes
                    ventas_ids.append(venta_id)
                    licencia = self.env['comercializador.licencia'].search([('cliente_id','=',import_line.get('cliente_id'))])
                    
                    if licencia: #si existe la licencia actualizo los valores
                        licencia.importe_venta = importe
                        licencia.ultima_venta = datetime.strptime(import_line.get('fecha_venta').strip(), '%d/%m/%Y').date()
                    else: # si no existe creo nueva licencia
                        licencia_id = self.env['comercializador.licencia'].create({
                            'cliente_id': import_line.get('cliente_id'),
                            'ultima_venta': datetime.strptime(import_line.get('fecha_venta').strip(), '%d/%m/%Y').date(),
                            'producto_id': self.producto_id.id,
                            'importe_venta': importe,
                            'division_id': self.division_id.id,
                            'fin_subscripcion': datetime.strptime(import_line.get('fecha_venta').strip(), '%d/%m/%Y').date() + relativedelta(months=1),
                        }).id
                        
                else: # cuando el importe de la venta es mayor a 370
                    ventas_mayores.append(import_line)
        
        if len(ventas_mayores) > 0:
            procesar_venta_id = self.env['comercializador.procesar.ventas.wiz'].create({
                    'producto_id': self.producto_id.id,
                    'division_id': self.division_id.id,
                }).id
            ventas_procesar_ids = []
            for venta in ventas_mayores:
                print("VENTA CLIENTE IDDD >>", venta.get('cliente_id'))
                venta_procesar_id = self.env['comercializador.venta.procesar'].create({
                        'fecha_venta': datetime.strptime(venta.get('fecha_venta').strip(), '%d/%m/%Y').date(),
                        'nombre_cliente': venta.get('nombre_cliente'),
                        'numero_factura': venta.get('factura'),
                        'codigo_concepto_facturacion': venta.get('codigo_concepto_facturacion'),
                        'importe_venta':  float(venta.get('importe_venta')),
                        'procesar_venta_wiz_id': procesar_venta_id,
                        'id_del_cliente': venta.get('cliente_id'),
                    })
                if venta.get('cliente_id') != 0:# si existe el cliente creo las bd del wizard
                    clientes_creado = self.env['comercializador.cliente'].sudo().search([('nombre','=',venta.get('nombre_cliente'))])
                    for cl in clientes_creado:
                        if cl.base_datos:
                            bd_cliente_id = self.env['comercializador.venta.base.datos'].create({
                                'nombre_base_datos': cl.base_datos,
                                'venta_procesar_id': venta_procesar_id.id,
                            }).id
                        else:
                            venta_procesar_id.pago_mayor = True
                ventas_procesar_ids.append(venta_procesar_id.id)
            
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'comercializador.procesar.ventas.wiz',
                'name': 'Procesar Ventas',
                'view_mode': 'form',
                'view_type': 'form',
                'views': [[False, 'form']],
                'target': 'new',
                'res_id': procesar_venta_id,
            }
            
        return {
                'type': 'ir.actions.act_window',
                'res_model': 'comercializador.venta',
                'name': 'Ventas',
                'view_mode': 'tree,form',
                'view_type': 'tree',
                'target': 'main', # Target
            }
            
        """ return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'type': 'success',
                    'title': _("Importando ventas"),
                    'message': _(f"Ventas importadas correctamente. Se crearon {len(new_clients_ids)} clientes nuevos. Se importaron {len(ventas_ids)} nuevas ventas"),
                    'next': {'type': 'ir.actions.act_window_close'},
                }
            } """
            

            